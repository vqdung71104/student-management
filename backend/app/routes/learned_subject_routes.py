from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import and_
from typing import List
import openpyxl
from io import BytesIO
from app.db.database import get_db
from app.models.__init__ import LearnedSubject, Subject, Student, SemesterGPA
from app.schemas.learned_subject_schema import (
    LearnedSubjectCreate,
    LearnedSubjectUpdate,
    LearnedSubjectResponse,
    LearnedSubjectSimpleCreate,
)

router = APIRouter(prefix="/learned-subjects", tags=["Learned Subjects"])

# 🔹 Hàm tính letter grade to score (thang 4.0)
def letter_grade_to_score(letter_grade: str) -> float:
    grade_map = {
        "F": 0.0,
        "D": 1.0,
        "D+": 1.5,
        "C": 2.0,
        "C+": 2.5,
        "B": 3.0,
        "B+": 3.5,
        "A": 4.0,
        "A+": 4.0
    }
    return grade_map.get(letter_grade, 0.0)

# 🔹 Hàm update semester GPA
def update_semester_gpa(student_id: int, semester: str, db: Session):
    # Lấy tất cả learned subjects của student trong semester này
    learned_subjects = db.query(LearnedSubject).filter(
        and_(
            LearnedSubject.student_id == student_id,
            LearnedSubject.semester == semester
        )
    ).all()
    
    if not learned_subjects:
        # Xóa SemesterGPA nếu không còn môn nào trong học kỳ này
        semester_gpa_to_delete = db.query(SemesterGPA).filter(
            and_(
                SemesterGPA.student_id == student_id,
                SemesterGPA.semester == semester
            )
        ).first()
        if semester_gpa_to_delete:
            db.delete(semester_gpa_to_delete)
        return
    
    total_credits = 0
    total_grade_points = 0.0
    
    for ls in learned_subjects:
        credits = ls.credits
        score = letter_grade_to_score(ls.letter_grade)
        total_credits += credits
        total_grade_points += credits * score
    
    gpa = total_grade_points / total_credits if total_credits > 0 else 0.0
    
    # Tìm hoặc tạo semester GPA
    semester_gpa = db.query(SemesterGPA).filter(
        and_(
            SemesterGPA.student_id == student_id,
            SemesterGPA.semester == semester
        )
    ).first()
    
    if semester_gpa:
        semester_gpa.gpa = gpa
        semester_gpa.total_credits = total_credits
    else:
        semester_gpa = SemesterGPA(
            student_id=student_id,
            semester=semester,
            gpa=gpa,
            total_credits=total_credits
        )
        db.add(semester_gpa)

# 🔹 Hàm update student stats
def update_student_stats(student_id: int, db: Session):
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        return
    
    # Lấy tất cả learned subjects của student
    learned_subjects = db.query(LearnedSubject).filter(LearnedSubject.student_id == student_id).all()
    
    # Reset counters
    failed_subjects_number = 0
    study_subjects_number = 0
    total_failed_credits = 0
    total_learned_credits = 0
    
    for ls in learned_subjects:
        if ls.letter_grade == "F":
            failed_subjects_number += 1
            total_failed_credits += ls.credits
        else:
            study_subjects_number += 1
            total_learned_credits += ls.credits
    
    # Update student fields
    student.failed_subjects_number = failed_subjects_number
    student.study_subjects_number = study_subjects_number
    student.total_failed_credits = total_failed_credits
    student.total_learned_credits = total_learned_credits
    
    # Calculate CPA from all semester GPAs
    semester_gpas = db.query(SemesterGPA).filter(SemesterGPA.student_id == student_id).all()
    total_credits = sum(sgpa.total_credits for sgpa in semester_gpas)
    total_grade_points = sum(sgpa.gpa * sgpa.total_credits for sgpa in semester_gpas)
    print("total_credits:", total_credits, "total_grade_points:", total_grade_points)
    if(total_learned_credits==0):
        student.cpa=0.0
    else:
        student.cpa = total_grade_points / total_learned_credits if total_credits > 0 else 0.0
    
    # Update warning level
    old_warning = student.warning_level
    if total_failed_credits >= 27:
        student.warning_level = "Cảnh báo mức 3"
    elif total_failed_credits >= 16:
        student.warning_level = "Cảnh báo mức 2"
    elif total_failed_credits >= 8:
        student.warning_level = "Cảnh báo mức 1"
    else:
        student.warning_level = "Cảnh báo mức 0"
    
    # Update level 3 warning counter
    if old_warning != "Cảnh báo mức 3" and student.warning_level == "Cảnh báo mức 3":
        student.level_3_warning_number += 1
    
    # Update year level
    total_learned = total_learned_credits
    if total_learned < 32:
        student.year_level = "Năm 1"
    elif total_learned < 64:
        student.year_level = "Năm 2"
    elif total_learned < 96:
        student.year_level = "Năm 3"
    elif total_learned < 128:
        student.year_level = "Năm 4"
    else:
        student.year_level = "Năm 5"

# 🔹 CRUD ROUTES

@router.get("/", response_model=list[LearnedSubjectResponse])
def get_all_learned_subjects(db: Session = Depends(get_db)):
    learned_subjects = db.query(LearnedSubject).all()
    return learned_subjects

@router.get("/{learned_subject_id}", response_model=LearnedSubjectResponse)
def get_learned_subject(learned_subject_id: int, db: Session = Depends(get_db)):
    learned_subject = db.query(LearnedSubject).filter(LearnedSubject.id == learned_subject_id).first()
    if not learned_subject:
        raise HTTPException(status_code=404, detail="Learned subject not found")
    return learned_subject

@router.get("/student/{student_id}", response_model=list[LearnedSubjectResponse])
def get_learned_subjects_by_student(student_id: int, db: Session = Depends(get_db)):
    learned_subjects = db.query(LearnedSubject).filter(LearnedSubject.student_id == student_id).all()
    return learned_subjects

@router.get("/student/{student_id}/semester/{semester}", response_model=list[LearnedSubjectResponse])
def get_learned_subjects_by_student_and_semester(
    student_id: int, 
    semester: str, 
    db: Session = Depends(get_db)
):
    learned_subjects = db.query(LearnedSubject).filter(
        and_(
            LearnedSubject.student_id == student_id,
            LearnedSubject.semester == semester
        )
    ).all()
    return learned_subjects

@router.post("/", response_model=LearnedSubjectResponse)
def create_learned_subject(learned_subject: LearnedSubjectCreate, db: Session = Depends(get_db)):
    # Get subject info
    subject = db.query(Subject).filter(Subject.id == learned_subject.subject_id).first()
    if not subject:
        raise HTTPException(status_code=404, detail="Subject not found")
    
    # Create learned subject
    db_learned_subject = LearnedSubject(
        subject_name=subject.subject_name,
        credits=subject.credits,
        letter_grade=learned_subject.letter_grade,
        semester=learned_subject.semester,
        student_id=learned_subject.student_id,
        subject_id=learned_subject.subject_id
    )
    
    db.add(db_learned_subject)
    db.commit()
    db.refresh(db_learned_subject)
    
    #    AUTO-CALCULATE GPA & STUDENT STATS
    update_semester_gpa(db_learned_subject.student_id, db_learned_subject.semester, db)
    update_student_stats(db_learned_subject.student_id, db)
    db.commit()
    
    return db_learned_subject

@router.put("/{learned_subject_id}", response_model=LearnedSubjectResponse)
def update_learned_subject(
    learned_subject_id: int, 
    learned_subject: LearnedSubjectUpdate, 
    db: Session = Depends(get_db)
):
    db_learned_subject = db.query(LearnedSubject).filter(LearnedSubject.id == learned_subject_id).first()
    if not db_learned_subject:
        raise HTTPException(status_code=404, detail="Learned subject not found")
    
    # Update fields
    for field, value in learned_subject.model_dump(exclude_unset=True).items():
        setattr(db_learned_subject, field, value)
    
    db.commit()
    db.refresh(db_learned_subject)
    
    #    AUTO-CALCULATE GPA & STUDENT STATS
    update_semester_gpa(db_learned_subject.student_id, db_learned_subject.semester, db)
    db.flush()  #    Flush to ensure semester_gpa is in session before calculating CPA
    update_student_stats(db_learned_subject.student_id, db)
    db.commit()
    
    return db_learned_subject

@router.delete("/{learned_subject_id}")
def delete_learned_subject(learned_subject_id: int, db: Session = Depends(get_db)):
    db_learned_subject = db.query(LearnedSubject).filter(LearnedSubject.id == learned_subject_id).first()
    if not db_learned_subject:
        raise HTTPException(status_code=404, detail="Learned subject not found")
    
    student_id = db_learned_subject.student_id
    semester = db_learned_subject.semester
    
    db.delete(db_learned_subject)
    db.commit()
    
    #    AUTO-CALCULATE GPA & STUDENT STATS after deletion
    # Cập nhật toàn bộ semester GPA của student (không chỉ semester bị xóa)
    all_semesters = db.query(LearnedSubject.semester).filter(
        LearnedSubject.student_id == student_id
    ).distinct().all()
    
    # Update GPA cho từng semester còn lại
    for (sem,) in all_semesters:
        update_semester_gpa(student_id, sem, db)
    
    # Xử lý riêng semester của môn vừa xóa (có thể đã không còn môn nào)
    update_semester_gpa(student_id, semester, db)
    
    db.flush()
    update_student_stats(student_id, db)
    db.commit()
    
    return {"message": "Learned subject deleted successfully"}


# 🔹 API: Sinh viên tự thêm môn học đã học
@router.post("/create-new-learned-subject")
def create_new_learned_subject(
    data: LearnedSubjectSimpleCreate,
    db: Session = Depends(get_db)
):
    # 1. Kiểm tra student tồn tại (data.student_id là students.id - INTEGER)
    student = db.query(Student).filter(Student.id == data.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail=f"Không tìm thấy sinh viên với ID {data.student_id}")
    
    # 2. Tra cứu subject từ mã HP (data.subject_id là mã HP string như "IT3080")
    subject = db.query(Subject).filter(Subject.subject_id == data.subject_id).first()
    if not subject:
        raise HTTPException(status_code=404, detail=f"Không tìm thấy môn học với mã HP {data.subject_id}")

    # 3. Kiểm tra trùng lặp theo mã học phần (bất kể học kỳ)
    existing = db.query(LearnedSubject).filter(
        and_(
            LearnedSubject.student_id == data.student_id,
            LearnedSubject.subject_id == subject.id,
        )
    ).first()

    old_semester = None
    if existing:
        # Lưu lại học kỳ cũ để cập nhật GPA sau khi xóa
        old_semester = existing.semester
        db.delete(existing)
        db.flush()  # Xóa khỏi session trước khi thêm mới

    # 4. Tạo LearnedSubject với student_id và subject.id (INTEGER FK)
    new_learned_subject = LearnedSubject(
        subject_name=subject.subject_name,
        credits=subject.credits,
        letter_grade=data.letter_grade,
        semester=data.semester,
        student_id=student.id,  # INTEGER: students.id
        subject_id=subject.id   # INTEGER: subjects.id
    )
    
    db.add(new_learned_subject)
    db.commit()
    db.refresh(new_learned_subject)
    
    # 6.    AUTO-CALCULATE GPA & STUDENT STATS
    # Nếu học kỳ cũ khác học kỳ mới thì cập nhật GPA cho cả 2 học kỳ
    if old_semester and old_semester != data.semester:
        update_semester_gpa(student.id, old_semester, db)
    update_semester_gpa(student.id, data.semester, db)
    db.flush()  #    Flush to ensure semester_gpa is in session before calculating CPA
    update_student_stats(student.id, db)
    db.commit()
    
    action = "Cập nhật" if old_semester else "Thêm"
    return {
        "message": f"{action} môn học thành công",
        "learned_subject": {
            "id": new_learned_subject.id,
            "subject_code": subject.subject_id,  # Trả về mã HP để hiển thị
            "subject_name": subject.subject_name,
            "credits": subject.credits,
            "letter_grade": data.letter_grade,
            "semester": data.semester
        }
    }

    


# 🔹 API: Upload điểm từ file Excel CTT HUST
@router.post("/upload-grades-excel")
async def upload_grades_excel(
    student_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    API upload file điểm Excel từ CTT HUST
    - Nhận student_id (MSSV string như "20225818") → tìm students.id
    - Parse Excel: Học kỳ, Mã HP (subject_id string), Tên HP, Tín chỉ, Điểm HP (chữ)
    - Mỗi dòng: tìm subjects.id từ mã HP, tạo LearnedSubject với FK integer
    - Xử lý giống POST "/" nhưng nhận student_id dạng string và parse từ Excel
    """
    try:
        # 1. Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Chỉ chấp nhận file Excel (.xlsx, .xls)")
        
        # 2. Tìm Student theo student_id (MSSV string) → lấy students.id (integer)
        student = db.query(Student).filter(Student.student_id == student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail=f"Không tìm thấy sinh viên với MSSV {student_id}")
        
        # 3. Read Excel file
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents))
        sheet = workbook.active
        
        # 4. Find header row
        header_row_idx = None
        required_fields = ['học kỳ', 'mã hp', 'điểm hp']
        
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            if row and all(any(field in str(cell).lower() for cell in row if cell) for field in required_fields):
                header_row_idx = idx
                break
        
        if not header_row_idx:
            raise HTTPException(
                status_code=400,
                detail="Không tìm thấy header với các cột bắt buộc: Học kỳ, Mã HP, Điểm HP"
            )
        
        # 5. Get headers
        headers = [str(cell).strip().lower() if cell else '' for cell in sheet[header_row_idx]]
        
        # Find column indices
        semester_idx = next((i for i, h in enumerate(headers) if 'học kỳ' in h or 'hoc ky' in h), None)
        subject_code_idx = next((i for i, h in enumerate(headers) if 'mã hp' in h), None)
        grade_idx = next((i for i, h in enumerate(headers) if 'điểm hp' in h or 'diem hp' in h), None)
        
        if semester_idx is None or subject_code_idx is None or grade_idx is None:
            raise HTTPException(
                status_code=400,
                detail="Không tìm thấy các cột bắt buộc: Học kỳ, Mã HP, Điểm HP"
            )
        
        # 6. Parse data rows
        created_count = 0
        skipped_count = 0
        errors = []
        valid_grades = ["A+", "A", "B+", "B", "C+", "C", "D+", "D", "F"]
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            # Skip empty rows
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            try:
                semester = str(row[semester_idx]).strip() if row[semester_idx] else ''
                subject_code = str(row[subject_code_idx]).strip() if row[subject_code_idx] else ''
                letter_grade = str(row[grade_idx]).strip() if row[grade_idx] else ''
                
                # Validate data
                if not semester or not subject_code or not letter_grade:
                    skipped_count += 1
                    continue
                
                if letter_grade not in valid_grades:
                    errors.append(f"Dòng {row_idx}: Điểm '{letter_grade}' không hợp lệ")
                    skipped_count += 1
                    continue
                
                # Find subject
                subject = db.query(Subject).filter(Subject.subject_id == subject_code).first()
                if not subject:
                    errors.append(f"Dòng {row_idx}: Không tìm thấy môn học với mã HP '{subject_code}'")
                    skipped_count += 1
                    continue
                
                # Check duplicate
                existing = db.query(LearnedSubject).filter(
                    and_(
                        LearnedSubject.student_id == student.id,
                        LearnedSubject.subject_id == subject.id,
                        LearnedSubject.semester == semester
                    )
                ).first()
                
                if existing:
                    skipped_count += 1
                    continue
                
                # Create learned subject
                new_learned_subject = LearnedSubject(
                    subject_name=subject.subject_name,
                    credits=subject.credits,
                    letter_grade=letter_grade,
                    semester=semester,
                    student_id=student.id,
                    subject_id=subject.id
                )
                
                db.add(new_learned_subject)
                created_count += 1
                
            except Exception as e:
                errors.append(f"Dòng {row_idx}: {str(e)}")
                skipped_count += 1
                continue
        
        # 7. Commit and update stats
        if created_count > 0:
            db.commit()
            
            # Update GPA for all affected semesters
            semesters = db.query(LearnedSubject.semester).filter(
                LearnedSubject.student_id == student.id
            ).distinct().all()
            
            for (semester,) in semesters:
                update_semester_gpa(student.id, semester, db)
            
            db.flush()  #    Flush to ensure all semester_gpas are in session before calculating CPA
            update_student_stats(student.id, db)
            db.commit()
        
        return {
            "message": "Upload điểm thành công",
            "created": created_count,
            "skipped": skipped_count,
            "errors": errors[:10] if errors else None  # Limit to first 10 errors
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Lỗi khi upload file: {str(e)}"
        )

