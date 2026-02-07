"""
Excel Export Service for Schedule Combinations
Generates formatted Excel files from chatbot schedule suggestions
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Dict, List, Any


class ExcelExportService:
    """Service to export schedule combinations to Excel format"""
    
    def __init__(self):
        # Define color scheme
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_excel(
        self,
        combination: Dict[str, Any],
        student_info: Dict[str, Any] = None
    ) -> BytesIO:
        """
        Generate Excel file from schedule combination data
        
        Args:
            combination: Schedule combination data from chatbot
            student_info: Optional student information (semester, CPA) - not used
        
        Returns:
            BytesIO object containing Excel file
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create only Class Details sheet (no overview sheet)
        self._create_class_details_sheet(wb, combination)
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    def _create_class_details_sheet(
        self,
        wb: Workbook,
        combination: Dict[str, Any]
    ):
        """Create class details sheet with full schedule table"""
        # Use simple sheet name: "Phương án X"
        combination_id = combination.get('combination_id', 1)
        ws = wb.create_sheet(f"Phương án {combination_id}", 0)
        
        # Headers
        headers = [
            "Mã lớp",
            "Tên lớp",
            "Thời gian",
            "Ngày học",
            "Tuần học",
            "Phòng",
            "Giáo viên",
            "Ghi chú"
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Write class data
        classes = combination.get('classes', [])
        for row_num, cls in enumerate(classes, 2):
            # Class ID
            cell = ws.cell(row=row_num, column=1)
            cell.value = cls.get('class_id', '-')
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
            
            # Class name
            cell = ws.cell(row=row_num, column=2)
            cell.value = cls.get('class_name', cls.get('subject_name', '-'))
            cell.border = self.border
            
            # Time
            cell = ws.cell(row=row_num, column=3)
            start_time = cls.get('study_time_start', '')
            end_time = cls.get('study_time_end', '')
            cell.value = f"{start_time} - {end_time}" if start_time and end_time else '-'
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
            
            # Study date
            cell = ws.cell(row=row_num, column=4)
            cell.value = cls.get('study_date', '-')
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
            
            # Study weeks
            cell = ws.cell(row=row_num, column=5)
            study_week = cls.get('study_week', [])
            if isinstance(study_week, list) and study_week:
                cell.value = ', '.join(map(str, study_week))
            else:
                cell.value = '-'
            cell.border = self.border
            
            # Classroom
            cell = ws.cell(row=row_num, column=6)
            cell.value = cls.get('classroom', '-')
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
            
            # Teacher
            cell = ws.cell(row=row_num, column=7)
            cell.value = cls.get('teacher_name', '-')
            cell.border = self.border
            
            # Note
            cell = ws.cell(row=row_num, column=8)
            cell.value = cls.get('priority_reason', '-')
            cell.border = self.border
        
        # Auto-adjust column widths
        column_widths = {
            'A': 12,  # Mã lớp
            'B': 35,  # Tên lớp
            'C': 18,  # Thời gian
            'D': 12,  # Ngày học
            'E': 40,  # Tuần học
            'F': 12,  # Phòng
            'G': 25,  # Giáo viên
            'H': 30,  # Ghi chú
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
